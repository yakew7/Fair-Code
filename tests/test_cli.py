import builtins
import importlib.util
import io
import json
from pathlib import Path
import sys

import pandas as pd
import pytest

import faircode.cli as cli
from faircode.cli import main

REPO_ROOT = Path(__file__).resolve().parent.parent
SMALL_AUDIT = REPO_ROOT / "German Credit Lending" / "audit.yaml"

requires_openpyxl = pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="optional 'excel' extra not installed",
)
requires_scipy = pytest.mark.skipif(
    importlib.util.find_spec("scipy") is None,
    reason="optional 'proxy' extra not installed",
)


def test_profile_fail_under_returns_nonzero_and_explains_score(tmp_path, capsys):
    path = tmp_path / "skewed.csv"
    path.write_text("sex\n" + "M\n" * 80 + "F\n" * 20, encoding="utf-8")

    exit_code = main(["profile", str(path), "--fail-under", "90"])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "Representation score:" in captured.out
    assert "representation score 72/100 is below --fail-under 90" in captured.err


def test_profile_fail_under_keeps_json_output_machine_readable(tmp_path, capsys):
    path = tmp_path / "balanced.csv"
    path.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--json", "--fail-under", "90"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert json.loads(captured.out)["overall_score"] == 100
    assert captured.err == ""


def test_profile_fail_under_equal_threshold_returns_zero(tmp_path, capsys):
    path = tmp_path / "balanced.csv"
    path.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--fail-under", "100"])

    captured = capsys.readouterr()

    assert exit_code == 0
    assert "Representation score: 100/100" in captured.out
    assert captured.err == ""


def test_compare_fail_on_drift_returns_nonzero_and_explains(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\n" + "M\n" * 50 + "F\n" * 50, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\n" + "M\n" * 90 + "F\n" * 10, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--fail-on-drift"])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "representation drift detected" in captured.err
    assert "--fail-on-drift" in captured.err


def test_compare_without_fail_on_drift_still_returns_zero(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\n" + "M\n" * 50 + "F\n" * 50, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\n" + "M\n" * 90 + "F\n" * 10, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b)])

    assert exit_code == 0


def test_compare_fail_on_drift_returns_zero_when_stable(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--fail-on-drift"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert captured.err == ""


def test_compare_applies_map_override_to_both_datasets(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("gndr\n" + "M\n" * 8 + "F\n" * 2, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("gndr\n" + "M\n" * 5 + "F\n" * 5, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--json", "--map", "gndr=sex"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert [d["name"] for d in result["dimensions"]] == ["gndr"]
    assert [d["kind"] for d in result["dimensions"]] == ["sex"]


def test_compare_without_map_leaves_column_generically_categorical(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("gndr\n" + "M\n" * 8 + "F\n" * 2, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("gndr\n" + "M\n" * 5 + "F\n" * 5, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert [d["kind"] for d in result["dimensions"]] == ["categorical"]


def test_profile_map_unknown_column_exits_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--map", "nonexistent_col=race"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "--map column(s) not found in the dataset: nonexistent_col" in captured.err


def test_compare_map_unknown_column_exits_2_with_clean_error(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["compare", str(path_a), str(path_b), "--map", "nonexistent_col=race"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "--map column(s) not found in the dataset: nonexistent_col" in captured.err


def test_compare_map_column_present_in_only_one_dataset_is_accepted(tmp_path):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex,extra\nM,1\nF,2\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--map", "extra=ignore", "--json"])

    assert exit_code == 0


def test_map_without_equals_sign_exits_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--map", "sex_no_equals"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "invalid --map 'sex_no_equals', expected COL=KIND" in captured.err


def test_map_with_invalid_kind_exits_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--map", "sex=not_a_real_kind"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "invalid --map kind 'not_a_real_kind' for column 'sex'" in captured.err


def test_profile_missing_file_exits_2_with_clean_error(tmp_path, capsys):
    missing = tmp_path / "does_not_exist.csv"

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(missing)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert f"error: file not found: {missing}" in captured.err


def test_profile_reads_csv_from_stdin(monkeypatch, capsys):
    monkeypatch.setattr("sys.stdin", io.StringIO("sex\nM\nF\nM\nF\n"))

    exit_code = main(["profile", "-", "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert result["n_rows"] == 4
    assert result["dimensions"][0]["name"] == "sex"


def test_profile_reads_tsv_from_stdin_via_delimiter_sniffing(monkeypatch, capsys):
    monkeypatch.setattr("sys.stdin", io.StringIO("sex\tage\nM\t30\nF\t40\n"))

    exit_code = main(["profile", "-", "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert result["n_cols"] == 2


def test_compare_reads_one_side_from_stdin(tmp_path, monkeypatch, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")
    monkeypatch.setattr("sys.stdin", io.StringIO("sex\nM\nM\nM\nF\n"))

    exit_code = main(["compare", str(path_a), "-", "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert result["a"]["n_rows"] == 4
    assert result["b"]["n_rows"] == 4


def test_compare_both_sides_from_stdin_returns_2_with_clean_error(capsys):
    exit_code = main(["compare", "-", "-"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--compare can't read both datasets from stdin" in captured.err


def test_profile_read_table_runtime_error_exits_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.parquet"
    path.write_text("not a real parquet file", encoding="utf-8")

    def raise_runtime(_path):
        raise RuntimeError("reading .parquet files requires the 'pyarrow' package")

    monkeypatch.setattr(cli, "read_table", raise_runtime)

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "error: reading .parquet files requires the 'pyarrow' package" in captured.err


def test_profile_read_table_generic_exception_exits_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    def raise_generic(_path):
        raise ValueError("boom")

    monkeypatch.setattr(cli, "read_table", raise_generic)

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert f"error: could not read dataset {path}: boom" in captured.err


def test_profile_malformed_cross_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--cross", "onlyonecolumn"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--cross expects two column names: COLA,COLB" in captured.err


def test_profile_cross_same_column_twice_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--cross", "sex,sex"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--cross needs two different columns, got 'sex' twice" in captured.err


def test_profile_cross_unknown_column_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex,race\n" + "M,A\nF,B\n" * 25, encoding="utf-8")

    exit_code = main(["profile", str(path), "--cross", "sex,raace"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "cross column(s) don't match any profiled dimension: raace" in captured.err


def test_profile_reference_missing_required_columns_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    ref_path = tmp_path / "ref.csv"
    ref_path.write_text("nothing,relevant\n1,2\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--reference", str(ref_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "reference needs column, group, and share columns" in captured.err


def test_profile_reference_with_no_matching_dimension_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nM\nF\n", encoding="utf-8")
    ref_path = tmp_path / "ref.csv"
    ref_path.write_text("column,group,share\ngendr,M,0.5\ngendr,F,0.5\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--reference", str(ref_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "reference file's column(s) don't match any profiled dimension: gendr" in captured.err


def test_profile_proxy_hints_runtime_error_returns_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    def raise_runtime(_df, _dimensions, **_kwargs):
        raise RuntimeError("proxy hints need scipy (install with: pip install faircode[proxy])")

    monkeypatch.setattr(cli, "proxy_hints", raise_runtime)

    exit_code = main(["profile", str(path), "--proxy-hints"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "error: proxy hints need scipy" in captured.err


@requires_scipy
def test_proxy_hints_with_flags_a_dropped_column(tmp_path, capsys):
    path = tmp_path / "dropped.csv"
    held_path = tmp_path / "full.csv"
    zip_code = (["111"] * 100 + ["222"] * 100)
    race = (["A"] * 100 + ["B"] * 100)  # perfectly aligned with zip_code
    path.write_text("zip_code\n" + "\n".join(zip_code), encoding="utf-8")
    held_path.write_text("zip_code,race\n" +
                         "\n".join(f"{z},{r}" for z, r in zip(zip_code, race)),
                         encoding="utf-8")

    exit_code = main(["profile", str(path), "--proxy-hints",
                      "--proxy-hints-with", f"{held_path}=race", "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    pair = next(h for h in result["proxy_hints"] if {h["a"], h["b"]} == {"zip_code", "race"})
    assert pair["p_value"] < 0.05


def test_proxy_hints_with_malformed_spec_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--proxy-hints", "--proxy-hints-with", "noequalssign"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "invalid --proxy-hints-with 'noequalssign'" in captured.err


def test_proxy_hints_with_unknown_column_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    held_path = tmp_path / "b.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    held_path.write_text("other\nx\ny\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--proxy-hints",
              "--proxy-hints-with", f"{held_path}=race"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert f"column 'race' not found in {held_path}" in captured.err


def test_proxy_hints_with_column_collision_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    held_path = tmp_path / "b.csv"
    path.write_text("sex,race\nM,A\nF,B\n", encoding="utf-8")
    held_path.write_text("race\nX\nY\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--proxy-hints",
              "--proxy-hints-with", f"{held_path}=race"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "column 'race' already exists in the profiled dataset" in captured.err


def test_proxy_hints_with_row_count_mismatch_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    held_path = tmp_path / "b.csv"
    path.write_text("sex\nM\nF\nM\n", encoding="utf-8")
    held_path.write_text("race\nA\nB\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--proxy-hints",
              "--proxy-hints-with", f"{held_path}=race"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "rows must align 1:1" in captured.err


@requires_openpyxl
def test_proxy_hints_with_xlsx_reports_ignored_sheets(tmp_path, capsys):
    import openpyxl

    path = tmp_path / "a.csv"
    path.write_text("sex\n" + "M\n" * 50 + "F\n" * 50, encoding="utf-8")

    held_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "Data"
    first.append(["race"])
    for _ in range(50):
        first.append(["A"])
    for _ in range(50):
        first.append(["B"])
    wb.create_sheet("Notes")
    wb.save(held_path)

    exit_code = main(["profile", str(path), "--proxy-hints",
                      "--proxy-hints-with", f"{held_path}=race"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert f"{held_path}: read sheet 'Data' - 1 other sheet(s) ignored." in captured.err


def test_proxy_hints_with_without_proxy_hints_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path),
                      "--proxy-hints-with", "/nonexistent/file.csv=race"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--proxy-hints-with needs --proxy-hints" in captured.err


requires_scipy = pytest.mark.skipif(
    importlib.util.find_spec("scipy") is None,
    reason="optional 'proxy' extra not installed",
)


@requires_scipy
def test_compare_proxy_hints_flags_a_real_proxy_in_both_datasets(tmp_path, capsys):
    # occupation is a perfect function of sex in both files -> maximal
    # association, so both A and B should surface the same proxy pair.
    rows = ["sex,occupation"] + [
        f"{'male' if i % 2 == 0 else 'female'},{'engineer' if i % 2 == 0 else 'nurse'}"
        for i in range(100)
    ]
    path_a = tmp_path / "a.csv"
    path_a.write_text("\n".join(rows), encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("\n".join(rows), encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--proxy-hints", "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    for key in ("proxy_hints_a", "proxy_hints_b"):
        pair = next(h for h in result[key] if {h["a"], h["b"]} == {"sex", "occupation"})
        assert pair["p_value"] < 0.05


def test_compare_proxy_hints_runtime_error_returns_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")

    def raise_runtime(_df, _dimensions, **_kwargs):
        raise RuntimeError("proxy hints need scipy (install with: pip install faircode[proxy])")

    monkeypatch.setattr(cli, "proxy_hints", raise_runtime)

    exit_code = main(["compare", str(path_a), str(path_b), "--proxy-hints"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "error: proxy hints need scipy" in captured.err


def test_profile_html_write_failure_returns_2_with_clean_error_not_a_traceback(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    bad_html_path = tmp_path / "no_such_dir" / "out.html"

    exit_code = main(["profile", str(path), "--html", str(bad_html_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert f"error: could not write HTML report to {bad_html_path}" in captured.err


def test_profile_html_write_success_reports_path(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    html_path = tmp_path / "out.html"

    exit_code = main(["profile", str(path), "--html", str(html_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert html_path.read_text(encoding="utf-8")
    assert f"HTML report written to {html_path}" in captured.err


def test_compare_html_write_failure_returns_2_with_clean_error_not_a_traceback(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")
    bad_html_path = tmp_path / "no_such_dir" / "out.html"

    exit_code = main(["compare", str(path_a), str(path_b), "--html", str(bad_html_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert f"error: could not write HTML report to {bad_html_path}" in captured.err


def test_compare_html_write_success_reports_path(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")
    html_path = tmp_path / "out.html"

    exit_code = main(["compare", str(path_a), str(path_b), "--html", str(html_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert html_path.read_text(encoding="utf-8")
    assert f"HTML report written to {html_path}" in captured.err


@requires_openpyxl
def test_profile_xlsx_reports_ignored_sheets(tmp_path, capsys):
    import openpyxl

    path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "Data"
    first.append(["sex"])
    first.append(["M"])
    first.append(["F"])
    wb.create_sheet("Notes")
    wb.create_sheet("Extra")
    wb.save(path)

    exit_code = main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "Read sheet 'Data' - 2 other sheet(s) ignored." in captured.err


@requires_openpyxl
def test_profile_xlsx_single_sheet_stays_silent(tmp_path, capsys):
    import openpyxl

    path = tmp_path / "single_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["sex"])
    ws.append(["M"])
    ws.append(["F"])
    wb.save(path)

    exit_code = main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "ignored" not in captured.err


def _make_multi_sheet_xlsx(path, sex_values):
    import openpyxl

    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "Data"
    first.append(["sex"])
    for value in sex_values:
        first.append([value])
    wb.create_sheet("Notes")
    wb.create_sheet("Extra")
    wb.save(path)


@requires_openpyxl
def test_compare_xlsx_reports_ignored_sheets_for_both_files(tmp_path, capsys):
    path_a = tmp_path / "a.xlsx"
    path_b = tmp_path / "b.xlsx"
    _make_multi_sheet_xlsx(path_a, ["M", "F"])
    _make_multi_sheet_xlsx(path_b, ["M", "M"])

    exit_code = main(["compare", str(path_a), str(path_b)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert f"{path_a}: read sheet 'Data' - 2 other sheet(s) ignored." in captured.err
    assert f"{path_b}: read sheet 'Data' - 2 other sheet(s) ignored." in captured.err


@requires_openpyxl
def test_compare_xlsx_single_sheet_stays_silent(tmp_path, capsys):
    import openpyxl

    path_a = tmp_path / "a.xlsx"
    path_b = tmp_path / "b.xlsx"
    for path, sex_values in ((path_a, ["M", "F"]), (path_b, ["M", "M"])):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["sex"])
        for value in sex_values:
            ws.append([value])
        wb.save(path)

    exit_code = main(["compare", str(path_a), str(path_b)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "ignored" not in captured.err


# ── Benchmark subcommand tests ────────────────────────────────────────────────

def test_cli_benchmark_import_error_message(monkeypatch, capsys):
    """Lines 245-249: Catch ImportError and emit the optional install guidance."""
    real_import = builtins.__import__

    def mock_import(name, *args, **kwargs):
        if "benchmark" in name:
            raise ImportError("No module named 'sklearn'")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", mock_import)
    monkeypatch.delitem(sys.modules, "faircode.benchmark", raising=False)

    exit_code = main(["benchmark"])
    assert exit_code == 2
    captured = capsys.readouterr()
    assert "error: the benchmark command needs scikit-learn and pyyaml" in captured.err
    assert "pip install faircode[benchmark]" in captured.err


def test_cli_benchmark_paper_drift_warning_on_overrides(monkeypatch, tmp_path, capsys):
    """Lines 253-263: Stderr warning when overriding frozen default resamples/permutations."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    dummy_fairness = pd.DataFrame([{"audit": "German Credit Lending", "metric": "dp"}])
    dummy_perf = pd.DataFrame([{"audit": "German Credit Lending", "metric": "auc"}])

    monkeypatch.setattr(
        "faircode.benchmark.run_benchmark",
        lambda **kwargs: (dummy_fairness, dummy_perf),
    )
    monkeypatch.setattr("faircode.benchmark.write_report", lambda *args, **kwargs: None)

    out_dir = str(tmp_path / "results")
    exit_code = main([
        "benchmark",
        "--n-resamples", "50",
        "--n-permutations", "50",
        "--out", out_dir,
        "--no-plots",
    ])

    assert exit_code == 0
    captured = capsys.readouterr()
    assert "warning: --n-resamples=50, --n-permutations=50 differs from the frozen paper-run default (2000)" in captured.err
    assert f"Ran 1 audit(s), wrote 1 fairness rows and 1 performance rows to {out_dir}/" in captured.err


def test_cli_benchmark_missing_matplotlib_returns_2_with_clean_error(monkeypatch, tmp_path, capsys):
    """write_report's lazy `from .figures import generate_figures` (only reached
    when make_plots=True) used to raise a raw ImportError traceback instead of
    the same clean error the top-level scikit-learn/pyyaml import already gets."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    dummy_fairness = pd.DataFrame([{"audit": "German Credit Lending", "metric": "dp"}])
    dummy_perf = pd.DataFrame([{"audit": "German Credit Lending", "metric": "auc"}])

    monkeypatch.setattr(
        "faircode.benchmark.run_benchmark",
        lambda **kwargs: (dummy_fairness, dummy_perf),
    )

    def raise_import_error(*args, **kwargs):
        raise ImportError("No module named 'matplotlib'")

    monkeypatch.setattr("faircode.benchmark.write_report", raise_import_error)

    exit_code = main(["benchmark", "--out", str(tmp_path / "results")])

    assert exit_code == 2
    captured = capsys.readouterr()
    assert "error: writing benchmark plots needs matplotlib" in captured.err
    assert "pip install faircode[benchmark]" in captured.err


def test_cli_benchmark_no_manifests_found_error(tmp_path, capsys):
    """Lines 271-273: Error exit path when no audit.yaml manifests are found in --root."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    empty_dir = tmp_path / "empty_root"
    empty_dir.mkdir()

    exit_code = main(["benchmark", "--root", str(empty_dir)])
    assert exit_code == 2
    captured = capsys.readouterr()
    assert f"error: no audit.yaml manifests found under {empty_dir}" in captured.err


def test_cli_benchmark_malformed_manifest_returns_2_with_clean_error(monkeypatch, tmp_path, capsys):
    """A malformed manifest or degenerate dataset used to crash `faircode
    benchmark` with a raw traceback (yaml.YAMLError/KeyError/sklearn
    ValueError) instead of the same clean CLI error every other anticipated
    failure in this subcommand already gets (#403)."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    def raise_value_error(**kwargs):
        raise ValueError("bad_audit/audit.yaml: 'target'")

    monkeypatch.setattr("faircode.benchmark.run_benchmark", raise_value_error)

    exit_code = main(["benchmark", "--out", str(tmp_path / "results")])

    assert exit_code == 2
    captured = capsys.readouterr()
    assert "error: bad_audit/audit.yaml: 'target'" in captured.err


@pytest.mark.skipif(not SMALL_AUDIT.is_file(), reason="German Credit Lending fixture not found")
def test_cli_benchmark_success_run(tmp_path, capsys):
    """Lines 274-283: Full benchmark execution against the German Credit Lending fixture."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    out_dir = tmp_path / "results"
    exit_code = main([
        "benchmark",
        str(SMALL_AUDIT),
        "--n-resamples", "5",
        "--n-permutations", "5",
        "--out", str(out_dir),
        "--no-plots",
    ])

    assert exit_code == 0
    captured = capsys.readouterr()
    assert "Ran 1 audit(s)" in captured.err
    assert f"to {out_dir}/" in captured.err
    assert (out_dir / "results_fairness.csv").is_file()
    assert (out_dir / "results_performance.csv").is_file()
    assert (out_dir / "summary.csv").is_file()
